from models import Customer, OldCustomer
from openpyxl import load_workbook
import time
from multiprocessing import Process
from openpyxl import load_workbook
from multiprocessing import Pool


class CustomerImport:
	newCustomerList = []
	matchingCustomers = 0
	currentCustomer = 1
	totalNewCustomers = 0
	existingCustomers = []


	def load_existing_customers(self):
		for existingCustomer in Customer.select():
			self.existingCustomers.append(existingCustomer)



	def get_new_customers(self):
		hasData = True
		row = 2
		wb = load_workbook(filename = 'job.xlsx')
		sheet_ranges = wb['Unit Job']

		print("Loading Customers From Spreadsheet")

		while hasData:
			newCustomer = OldCustomer()
			name = sheet_ranges['Q{0}'.format(row)].value
			newCustomer.parse_names(name)
			newCustomer.old_id = sheet_ranges['A{0}'.format(row)].value
			newCustomer.phone = sheet_ranges['R{0}'.format(row)].value
			newCustomer.mobile = sheet_ranges['S{0}'.format(row)].value
			newCustomer.email_address = sheet_ranges['T{0}'.format(row)].value
			newCustomer.line1 = sheet_ranges['U{0}'.format(row)].value
			newCustomer.line2 = sheet_ranges['V{0}'.format(row)].value
			newCustomer.line3 = sheet_ranges['W{0}'.format(row)].value
			newCustomer.city = sheet_ranges['X{0}'.format(row)].value
			newCustomer.postal_code = sheet_ranges['Y{0}'.format(row)].value
			newCustomer.correct_numbers()

			#print("Getting customer {0}. ID: {1}".format(newCustomer.display_name, newCustomer.old_id))

			newCustomer.save()


			nextRow = sheet_ranges['Q{0}'.format(row + 1)].value

			if nextRow is None:
				hasData = False
				self.totalNewCustomers = len(self.newCustomerList)
			row += 1



	def start_match_process(self):
		matchCount = 0
		procs = []
		pool = Pool(processes=3)
		print("GETTING MATCHES")
		results = pool.map(self.get_matches, OldCustomer.select())


		for result in results:
			for matchString in result:
				matchCount += 1
				with open("old_dupes.txt", "a") as myfile:
					myfile.write("\n\n--------------------------------------DUPE NUMBER {0} FOUND ---------------------------------------------------".format(matchCount))
					myfile.write(matchString)

		print("Total Matches: {0}".format(matchCount))




	def get_matches(self, customer):
		self.get_new_customers()
		matchStrings = [self.generate_match_string(customer, oldCustomer) for oldCustomer in OldCustomer.select() if customer.check_match(oldCustomer) and customer.old_id != oldCustomer.old_id]

		return(matchStrings)

	def generate_match_string(self, customer, existingCustomer):
		matchString = """   \n-----------OLD CUSTOMER----------")
								    \nDisplay Name: {0}"
								    \nTitle: {1}"
								    \nGiven Name: {2}"
								    \nFamily Name: {3}"
								    \nEmail Adress: {4}"
								    \nLim: {5}"
								    \nDisplay Name: {6}"
								    \n\n-----------Existing CUSTOMER----------")
								    \nDisplay Name: {7}"
								    \nTitle: {8}"
								    \nGiven Name: {9}"
								    \nFamily Name: {10}"
								    \nEmail Adress: {11}"
								    \nLine 1: {12}"
								    \nPostal Code: {13}""".format(customer.display_name, customer.title, customer.given_name, customer.family_name, customer.email_address, customer.line1, customer.postal_code, 
								    	existingCustomer.display_name, existingCustomer.title, existingCustomer.given_name, existingCustomer.family_name, existingCustomer.email_address, existingCustomer.line1, existingCustomer.postal_code)

		return matchString





if __name__ == '__main__':

	customerImport = CustomerImport()
	customerImport.get_new_customers()

	print ("\nnumber of customers: {0}".format(customerImport.totalNewCustomers))

	startTime = time.time()
	customerImport.start_match_process()
	endTime = time.time()

	print("Execution Time: {0}".format(endTime - startTime))

	matchingPercentage = customerImport.matchingCustomers / 100 * customerImport.totalNewCustomers

	print("\n\nThere were {0} matching customers, {1} % of the total".format(customerImport.matchingCustomers, matchingPercentage))

	print("COMPLETED")


