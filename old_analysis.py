from models import Customer
from openpyxl import load_workbook
import time
from datetime import datetime, timedelta
import json
import operator

hasData = True
row = 2
jobDates = []
jobDateDict = {}
dateTotals = {}
monthsList = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
yearList = [2010, 2011, 2012, 2013, 2014, 2015, 2016, 2017, 2018]
yearDict = {2010 : {'totalCount': 0, 'days' : {}}, 2011 : {'totalCount': 0, 'days' : {}}, 2012 : {'totalCount': 0, 'days' : {}}, 2013 : {'totalCount': 0, 'days' : {}}, 
2014: {'totalCount': 0, 'days' : {}}, 2015 : {'totalCount': 0, 'days' : {}}, 2016 : {'totalCount': 0, 'days' : {}}, 2017 : {'totalCount': 0, 'days' : {}}, 2018 : {'totalCount': 0, 'days' : {}}}
from openpyxl import load_workbook
wb = load_workbook(filename = 'job.xlsx')
sheet_ranges = wb['Unit Job']
matchingCustomers = 0
newJobs = None
engineerTotals = {}

while hasData:
	jobDateTime = datetime.strptime(sheet_ranges['Z{0}'.format(row)].value, '%d/%m/%Y') 
	jobDate = jobDateTime.date()
	engineer = sheet_ranges['B{0}'.format(row)].value

	if engineer in engineerTotals:
		engineerTotals[engineer] += 1
	else:
		engineerTotals[engineer] = 1




	print("Getting Date row {0}".format(row))

	jobDates.append(jobDateTime)
	if jobDate in jobDateDict:
		jobDateDict[jobDate] += 1
	else:
		jobDateDict[jobDate] = 1


	nextRow = sheet_ranges['Z{0}'.format(row + 1)].value

	if nextRow is None:
		hasData = False
	row += 1

with open('jobs_data.json') as data_file:    
    newJobs = json.load(data_file)

for job in newJobs:

	jobDateTime = datetime.strptime(job['created'], '%d/%m/%Y %H:%M')
	jobDate = jobDateTime.date()
	engineer = job['engineer']['first_name']

	if engineer in engineerTotals:
		engineerTotals[engineer] += 1
	else:
		engineerTotals[engineer] = 1

	jobDates.append(jobDateTime)

	if jobDate in jobDateDict:
		jobDateDict[jobDate] += 1
	else:
		jobDateDict[jobDate] = 1

for year in yearList:
	dateTotals[year] = {}

	for month in monthsList:
		dateTotals[year][month] = 0



for jobDate in jobDates:
	jobYear = jobDate.year
	jobMonth = jobDate.strftime("%B")
	dateTotals[jobYear][jobMonth] += 1

for key, val in dateTotals.items():
	totalJobs = 0
	for month, monthCount in val.items():
		totalJobs += monthCount

	monthAverage = totalJobs / 12

	dateTotals[key]["Average Jobs Per Month"] = int(round(monthAverage, 0))
	dateTotals[key]["Total Jobs"] = totalJobs

#print(jobDateDict)
#time.sleep(10)
#topJobDate = max(jobDateDict.items(), key=operator.itemgetter(1))[0]
#topJobDateCount = jobDateDict[topJobDate]

jobDateFile = open("jobdates.txt", "w")

topJobDate = None
topJobDateCount = 0
for date, count in jobDateDict.items():
	jobDateFile.write("\nDate: {0}Count: {1}".format(date, count))
	if count > topJobDateCount:
		topJobDateCount = count
		topJobDate = date

jobDateFile.close()

the_file = open("old_analysis.txt", "w")

yearJobCount = yearDict

for key, val in jobDateDict.items():
	jobCount = 0
	baseStartDate = datetime(datetime.today().year, 1, 1).date()
	baseCurrentDate = datetime.now().date()

	for year in yearList:
		startDate = baseStartDate.replace(year=year)
		currentDate = baseCurrentDate.replace(year=year)

		if key <= currentDate and key >= startDate:
			yearJobCount[key.year]['totalCount'] += val
			#print("Year {0} day {1}".format(key.year, key.day))

			daysDict = yearJobCount[key.year]['days']

			if key.day in daysDict:
				daysDict[key.day] += val
			else:
				daysDict[key.day] = val



the_file.write("There have been {0} jobs so far this year.".format(yearJobCount))

for key, val in dateTotals.items():
	the_file.write("\n\n\nJOB NUMBERS FOR {0}".format(key))
	for month, monthCount in val.items():
		the_file.write("\n{0}: {1}".format(month, monthCount))

the_file.write("\n\n\nENGINEER TOTALS")
for key, val in engineerTotals.items():
	the_file.write("\n{0}: {1}".format(key, val))

the_file.write("\n\nBUSIEST DAY")
the_file.write("\nThe Busiest day was {0}, with {1} jobs".format(topJobDate.strftime("%A %B %d %Y"), topJobDateCount))

the_file.write("\n\nCURRENT YEAR JOB COUNT COMPARISON")

for year, data in yearJobCount.items():
	if year == datetime.now().year:
		the_file.write("\n\nThere have been {0} jobs so far this year.".format(data['totalCount']))

		for thisDay, thisCount in sorted(data['days'].items()):
			the_file.write("\nOn the {0} of the month this year there were {1} jobs".format(thisCount, thisDay))
	else:
		the_file.write("\n\nThere were {0} jobs by this time in {1}".format(data['totalCount'], year))
		for day, count in data['days'].items():
			the_file.write("\nOn the {0} of the month there were {1} jobs".format(day, count))

the_file.close()